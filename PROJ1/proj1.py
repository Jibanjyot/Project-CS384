import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import BLACK, BLUE
from sendMail import *
import os
import pandas as pd 

roll_email_mapping={}
def getMailData():
    for roll,email in roll_email_mapping.items():
        for i in email:
            send_mail(roll,i)
    return     
def generateMarksheets(p_mark,n_mark):

    roll_name_mapping = {}
    with open('uploads/master_roll.csv') as roll_file:
        responses = csv.DictReader(roll_file)
        for row in responses:
            roll_name_mapping[row["roll"]] =row["name"]
            
    # print(roll_name_mapping)

    with open('uploads/responses.csv') as response_file:
        responses = csv.reader(response_file)
        response_master_list = []
        for row in responses:
            response_master_list.append(row)
            roll_email_mapping[row[6]] = [row[1],row[4]]
        # print(response_master_list)
    
    concise_sheet_header = response_master_list[0][:6].copy()
    master_answer_key = response_master_list[1]
    concise_sheet = [response_master_list[0]]
    # concise_sheet_header = response_master_list[0].copy()
    response_master_list = response_master_list[1:]
    # print(master_answer_key)
    exist_roll_mapping = [i[6] for i in response_master_list]
    for i in roll_name_mapping.keys():
        if i not in exist_roll_mapping:
            temp = master_answer_key
            temp = ["" for i in temp]
            temp[3] = roll_name_mapping[i]
            temp[6] = i
            response_master_list.append(temp)
            # print(temp)
    # print(exist_roll_mapping)
    
    for stud_detail in response_master_list:
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 16.89
        ws.column_dimensions['B'].width = 16.89
        ws.column_dimensions['C'].width = 16.89
        ws.column_dimensions['D'].width = 16.89
        ws.column_dimensions['E'].width = 16.89

        ws.row_dimensions[5].height = 22.2

        ws.merge_cells('A5:E5')
        top_left_cell = ws["A5"]
        top_left_cell.value = "Mark Sheet"
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        
        # for row in ws['A5:E5']:
        #     for cell in row:
        #         cell.border = Border(top=double, left=thin, right=thin, bottom=thin)
        #         print(cell)
        
        top_left_cell.font  = Font(b=True,name="Century", color=BLACK,underline="single")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")


        img = Image('IITPLOGO.PNG')
        img.width = 631.55905512
        img.height = 82.393700787
        ws.add_image(img,'A1')

        ws.cell(row=6,column=1).value = "Name:"
        ws.cell(row=6,column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=6,column=2).value = stud_detail[3]
        ws.cell(row=6,column=2).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=7,column=1).value = "Roll Number:"
        ws.cell(row=7,column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=7,column=2).value = stud_detail[6]
        ws.cell(row=7,column=2).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=6,column=4).value = "Exam:"
        ws.cell(row=6,column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=6,column=5).value = "quiz"
        ws.cell(row=6,column=5).font  = Font(b=True,name="Century", color=BLACK)

        ws.cell(row=9,column=2).value = "Right"
        ws.cell(row=9,column=2).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=9,column=3).value = "Wrong"
        ws.cell(row=9,column=3).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=9,column=4).value = "Not Attempt"
        ws.cell(row=9,column=4).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=9,column=5).value = "Max"
        ws.cell(row=9,column=5).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=10,column=1).value = "No."
        ws.cell(row=10,column=1).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=11,column=1).value = "Marking"
        ws.cell(row=11,column=1).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=12,column=1).value = "Total"
        ws.cell(row=12,column=1).font  = Font(b=True,name="Century", color=BLACK)

        ws.cell(row=15,column=1).value = "Student Ans"
        ws.cell(row=15,column=1).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=15,column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.cell(row=15,column=2).value = "Correct Ans"
        ws.cell(row=15,column=2).font  = Font(b=True,name="Century", color=BLACK)
        ws.cell(row=15,column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_number = 16
        col_number = 1
        tot_correct = 0
        tot_wrong = 0
        tot_not_attempted = 0
        for i in range(7,len(master_answer_key)):
            if row_number == 41:
                row_number = 15
                col_number = 4
                ws.cell(row=row_number,column=col_number).value = "Student Ans"
                ws.cell(row=row_number,column=col_number).font  = Font(b=True,name="Century", color=BLACK)
                ws.cell(row_number,column=col_number).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(row=row_number,column=col_number+1).value = "Correct Ans"
                ws.cell(row=row_number,column=col_number+1).font  = Font(b=True,name="Century", color=BLACK)
                ws.cell(row_number,column=col_number+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                row_number=row_number+1

             

            ws.cell(row_number,column=col_number).value = stud_detail[i]
            ws.cell(row_number,column=col_number).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.cell(row_number,column=col_number+1).value = master_answer_key[i]
            ws.cell(row_number,column=col_number+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.cell(row_number,column=col_number+1).font  = Font(b=True,name="Century", color="000000FF")
            if stud_detail[i] == "":
                tot_not_attempted = tot_not_attempted+1
            elif stud_detail[i] == master_answer_key[i]:
                tot_correct=tot_correct+1
                ws.cell(row_number,column=col_number).font  = Font(b=True,name="Century", color="00008000")
            elif stud_detail[i] != master_answer_key[i]:
                tot_wrong = tot_wrong+1
        
                ws.cell(row_number,column=col_number).font  = Font(b=True,name="Century", color="00FF0000")
            row_number = row_number+1
        
        #input to be taken
        positive_marks = p_mark
        negative_marks = n_mark
        total_positive_marks = tot_correct*positive_marks
        total_negative_marks = tot_wrong*negative_marks     
        total_score = total_positive_marks-total_negative_marks
        # print(total_positive_marks)
        tot_marks = (len(master_answer_key)-7)*positive_marks


        ws.cell(row=10,column=2).value = tot_correct
        ws.cell(row=10,column=2).font  = Font(color="00008000")
        ws.cell(row=10,column=3).value = tot_wrong
        ws.cell(row=10,column=3).font  = Font(color="00FF0000")
        ws.cell(row=10,column=4).value = tot_not_attempted
        ws.cell(row=11,column=2).value = positive_marks
        ws.cell(row=11,column=2).font  = Font(color="00008000")
        ws.cell(row=11,column=3).value = int("-"+str(negative_marks))
        ws.cell(row=11,column=3).font  = Font(color="00FF0000")
        ws.cell(row=11,column=4).value = 0
        ws.cell(row=12,column=2).value = total_positive_marks
        ws.cell(row=12,column=2).font  = Font(color="00008000")
        ws.cell(row=12,column=3).value = total_negative_marks
        ws.cell(row=12,column=3).font  = Font(color="00FF0000")
        ws.cell(row=12,column=5).value = str(total_score)+"/"+str(tot_marks)
        ws.cell(row=12,column=5).font  = Font(color="000000FF")

        

        for row in ws['A9:E12']:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # print(cell)

     


        for row in ws['A8:E58']:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # print(cell)
        






        wb_name = "marksheets/"+stud_detail[6]+'.xlsx'
        wb.save(wb_name)
        temp_list = stud_detail.copy()
        temp_list.insert(6,str(total_score)+"/"+str(tot_marks))
        temp_list.append([tot_correct,tot_wrong,tot_not_attempted])
        concise_sheet.append(temp_list)
        
        
    
    concise_sheet_header[2] = "Google_Score"
    concise_sheet_header.insert(6,"Score_After_Negative")
    for i in range(7,len(master_answer_key)):
        concise_sheet_header.append("unnamed: "+str(i))
    concise_sheet_header.append("statusAnswer")
    # with open("marksheets/concise.csv",'w',newline='') as f:
    #     writer = csv.writer(f)
    #     writer.writerow(concise_sheet_header)
    #     for row in concise_sheet:
    #         writer.writerow(row)
    #         print(row)


    return    



def generateconciseMarksheets(p_mark,n_mark):

    roll_name_mapping = {}
    with open('uploads/master_roll.csv') as roll_file:
        responses = csv.DictReader(roll_file)
        for row in responses:
            roll_name_mapping[row["roll"]] =row["name"]
    # print(roll_name_mapping)

    with open('uploads/responses.csv') as response_file:
        responses = csv.reader(response_file)
        response_master_list = []
        for row in responses:
            response_master_list.append(row)
        # print(response_master_list)
    
    concise_sheet_header = response_master_list[0][:6].copy()
    master_answer_key = response_master_list[1]
    concise_sheet = [response_master_list[0]]
    # concise_sheet_header = response_master_list[0].copy()
    response_master_list = response_master_list[1:]
    # print(master_answer_key)
    exist_roll_mapping = [i[6] for i in response_master_list]
    for i in roll_name_mapping.keys():
        if i not in exist_roll_mapping:
            temp = master_answer_key
            temp = ["" for i in temp]
            temp[3] = roll_name_mapping[i]
            temp[6] = i
            response_master_list.append(temp)
            # print(temp)
    # print(exist_roll_mapping)
    
    for stud_detail in response_master_list:
        row_number = 16
        col_number = 1
        tot_correct = 0
        tot_wrong = 0
        tot_not_attempted = 0
        for i in range(7,len(master_answer_key)):
            if row_number == 41:
                row_number = 15
                col_number = 4
                row_number=row_number+1
            if stud_detail[i] == "":
                tot_not_attempted = tot_not_attempted+1
            elif stud_detail[i] == master_answer_key[i]:
                tot_correct=tot_correct+1
            elif stud_detail[i] != master_answer_key[i]:
                tot_wrong = tot_wrong+1
            row_number = row_number+1
        
        #input to be taken
        positive_marks = p_mark
        negative_marks = n_mark
        total_positive_marks = tot_correct*positive_marks
        total_negative_marks = tot_wrong*negative_marks     
        total_score = total_positive_marks-total_negative_marks
        # print(total_positive_marks)
        tot_marks = (len(master_answer_key)-7)*positive_marks
        # if stud_detail[6] not in exist_roll_mapping:
        #     print(stud_detail[6])


        temp_list = stud_detail.copy()
        temp_list[2] = str(total_positive_marks)+"/"+str(tot_marks)
        
        temp_list.insert(6,str(total_score)+"/"+str(tot_marks))
        temp_list.append([tot_correct,tot_wrong,tot_not_attempted])
        if stud_detail[6] not in exist_roll_mapping:
            temp_list[2] = "Absent"
            temp_list[6] = "Absent"

        concise_sheet.append(temp_list)
        
        
    
    concise_sheet_header[2] = "Google_Score"
    concise_sheet_header.insert(6,"Score_After_Negative")
    for i in range(7,len(master_answer_key)+1):
        concise_sheet_header.append("unnamed: "+str(i))
    concise_sheet_header.append("statusAnswer")
    # with open("marksheets/concise.csv",'w',newline='') as f:
    #     writer = csv.writer(f)
    #     writer.writerow(concise_sheet_header)
    #     for row in concise_sheet:
    #         writer.writerow(row)
    #         # print(row)
    concise_sheet = concise_sheet[1:]
    df = pd.DataFrame(concise_sheet, columns = concise_sheet_header)
    df.to_csv("marksheets/concise.csv")

    return


