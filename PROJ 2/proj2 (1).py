import csv
from openpyxl import Workbook
import os
from fpdf import FPDF
import datetime

def func():
    nameRollMapping = {}
    nestedDict={}
    with open('sample_input/names-roll.csv','r') as namesRollFile:
        reader = csv.DictReader(namesRollFile)
        for row in reader:
            nameRollMapping[row['Roll']] = row['Name']
            nestedDict[row['Roll']]={}
    subjectMapping = {}
    with open('sample_input/subjects_master.csv','r') as subject_master:
        reader = csv.DictReader(subject_master)
        for row in reader:
            subjectMapping[row['subno']] = [row['subname'],row['crd'],row['ltp']]
    # print(subjectMapping)
    gradeswithsem ={}
    rolls = []
    with open('sample_input/grades.csv','r') as gradesFile:
        reader = csv.DictReader(gradesFile)
        for row in reader:
            if row['Roll'] in rolls:
                gradeswithsem[row['Roll']].append([row['Sem'],row['SubCode'],subjectMapping[row['SubCode']][0],subjectMapping[row['SubCode']][2],row['Credit'],row['Sub_Type'],row['Grade']])
            else:
                gradeswithsem[row['Roll']] = [[row['Sem'],row['SubCode'],subjectMapping[row['SubCode']][0],subjectMapping[row['SubCode']][2],row['Credit'],row['Sub_Type'],row['Grade']]]
                rolls.append(row['Roll'])
    # print(gradeswithsem)
    for key,val in gradeswithsem.items():
        nestedDict[key]['1']=[]
        nestedDict[key]['2']=[]
        nestedDict[key]['3']=[]
        nestedDict[key]['4']=[]
        nestedDict[key]['5']=[]
        nestedDict[key]['6']=[]
        nestedDict[key]['7']=[]
        nestedDict[key]['8']=[]
        nestedDict[key]['9']=[]
        nestedDict[key]['10']=[]
    for key,val in gradeswithsem.items():
        for data in val:
            nestedDict[key][data[0]].append(data)
    # print(nestedDict['0401CS01'])
                
    grade = {'AA': 10,"AB":9,'BB':8,' BB':8,'BC': 7, 'CC':6,'CD':5,'DD':4,'F':0,'I':0,'F*':0,'I*':0,'DD*':4}
    overalldata = {}
    for key in gradeswithsem:
        overalldata[key] = []
    # print(overalldata)
    for key in gradeswithsem:
        #semdata n rows 3 columns semno,semcred,spi,totcred,cpi
        semadded = []
        semdata1 = []
        
        tot_sem_cred = 0
        tot_cred = 0
        sem = 1
        numerator = 0
        semdata = [0,0,0,0,0]
        loopno = 1
        cpi = 0
        for item in gradeswithsem[key]:
            # print(loopno,"  ",len(gradeswithsem[key]))
            if sem == 9:
                break
            if int(item[0]) == sem:
                tot_sem_cred = tot_sem_cred + int(item[4])
                semdata[1] = tot_sem_cred
                numerator = numerator + int(item[4])*grade[item[6]]
            
            else:
                semdata[1] = tot_sem_cred
                spi = numerator/tot_sem_cred
                cpi = (cpi*tot_cred+spi*tot_sem_cred)/(tot_cred+tot_sem_cred)
                semdata[4] = round(cpi,2)
                tot_cred = tot_cred+tot_sem_cred
                semdata[3] = tot_cred
                spi = numerator/tot_sem_cred
                
                semdata[2] = round(spi,2)
                semdata[0] = sem
                semdata1.append(semdata)
                overalldata[key].append(semdata)
                nestedDict[key][str(sem)].append(semdata)
                sem = sem + 1
                semdata = [0,0,0,0,0]
                tot_sem_cred = 0
                numerator=0
                tot_sem_cred = tot_sem_cred + int(item[4])
                numerator = int(item[4])*grade[item[6]]
            if loopno == len(gradeswithsem[key]):
                # print("yes")
                semdata[1] = tot_sem_cred
                spi = numerator/tot_sem_cred
                cpi = (cpi*tot_cred+spi*tot_sem_cred)/(tot_cred+tot_sem_cred)
                semdata[4] = round(cpi,2)
                tot_cred = tot_cred+tot_sem_cred
                semdata[3] = tot_cred
                spi = numerator/tot_sem_cred
                semdata[2] = round(spi,2)
                semdata[0] = sem
                semdata1.append(semdata)
                overalldata[key].append(semdata)
                nestedDict[key][str(sem)].append(semdata)
            loopno = loopno+1
    # print(nestedDict)
    # print(nestedDict['0401CS01'],'1','a','2','a','r')
    generate_mtechphd_marksheet(nestedDict['0401CS01'],'1901ME29','Jibanjyoti Kalita','2020','Master of Technology','Mechanical Engineering')
    
    return

def generate_btech_marksheet(transcript_data,rollno,name,yoa,programme,course):
    transcript_data = transcript_data
    pdf = FPDF()

    pdf.add_page(format = 'A3')
    pdf.set_font("Times", size=10)
    line_height = pdf.font_size * 2.5
    # print(pdf.y)
    #----------------------------------------------------------#
    top = pdf.y
    xcoord = pdf.x
    pdf.cell(280, 388, '', 1, 1, 'C')
    pdf.y = top
    pdf.image("IITPLOGO2.PNG",x = xcoord+1, y = top+1, w = 275, h = 30)
    pdf.ln(30)

    #----------------------------------------------------------#
    stud_info = [["Roll No: ",rollno,"Name: ",name, "Year of Admission: ",yoa], 
                ["Programme: ",programme,"Course: ",course],"",""]
    line_height = 10
    pdf.x = 50
    top = pdf.y
    pdf.cell(215, 23, '', 1, 1, 'C')
    pdf.y = top

    for row in stud_info:
        pdf.x = 50
        col_no = 1
        for datum in row:
            if col_no == 1:
                col_width = 22
            if col_no == 2:
                col_width = 50
            if col_no == 3:
                col_width = 20
            if col_no == 4:
                col_width = 50
            if col_no == 5:
                col_width = 35
            if col_no == 6:
                col_width = 20
            pdf.multi_cell(col_width, line_height, datum, border=0, ln=3, max_line_height=pdf.font_size)
            col_no = col_no + 1
        pdf.ln(line_height)

        #----------------------------------------------------------#
    xcoord = pdf.x
    pdf.y = pdf.y-13
    ycoord = pdf.y
    offset = 3
    offset_y = 100
    for key in transcript_data:
        sem_table_list = []
        if key == "9":
            break
        if key == "4" or key == "7":
            pdf.y = pdf.y + offset_y
            ycoord = pdf.y
            pdf.x = xcoord
            offset = 3

        for subject_wise_data in transcript_data[key]:
            sem_table_list.append(subject_wise_data)
        # print(sem_table_list)
        # break
        pdf.x = pdf.x+offset
        pdf.set_font('Arial',"", 8)
        sem_name = "Semester "+str(key)
        pdf.set_font('Arial', 'BU', 6)
        pdf.cell(20, 10,sem_name, 0, 1, 'C')
        pdf.ln(1)
        for row in sem_table_list[:-1]:
            i = 1
            
            pdf.x = pdf.x+offset
            for datum in row[1:]:
                if i == 1:
                    cell_width = 9
                if i == 2:
                    cell_width = 40
                if i == 3:
                    cell_width = 9
                if i == 4:
                    cell_width = 9
                if i == 5:
                    cell_width = 9
                pdf.set_font('Arial', '', 5)
                
                pdf.multi_cell(cell_width, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                i=i+1
            pdf.ln(line_height)
        i=1
        # print(sem_table_list[-1])
        pdf.x = pdf.x+offset
        pdf.y = pdf.y+2
        pdf.set_font('Arial', 'B', 7)
        overalldata = "Credits Taken: {creditstaken}     Credits Cleared: {creditscleared}   SPI: {spi}  CPI: {cpi}".format(creditstaken=sem_table_list[-1][1],creditscleared = sem_table_list[-1][3],spi=sem_table_list[-1][2],cpi=sem_table_list[-1][4] )
        pdf.cell(77, 8, overalldata, 1, 1, 'C')
        pdf.ln(line_height)
        

        pdf.y = ycoord
        offset = offset + 90


    today = datetime.datetime.now()
    today = today.strftime("%d %b %Y, %H:%M")
    today = "Date Generated: "+str(today)
    pdf.x = 30
    pdf.y = 365
    pdf.set_font('Arial', '', 9)
    pdf.cell(20, 10,today, 0, 1, 'C')

    pdf.x = 120
    pdf.y = 365
    pdf.image("SEAL.PNG",x = 120, y = 360, w = 35, h = 35)


    pdf.x = 230
    pdf.y = 365
    pdf.cell(20, 10,"Assistant Registrar (Academic)", 0, 1, 'C')
    file_name = str(rollno)+".pdf"
    pdf.output(file_name)

def generate_mtechphd_marksheet(transcript_data,rollno,name,yoa,programme,course):
    transcript_data = transcript_data
    pdf = FPDF()

    pdf.add_page(format = 'A4')
    pdf.set_font("Times", size=10)
    line_height = pdf.font_size * 2.5
    # print(pdf.y)
    #----------------------------------------------------------#
    top = pdf.y
    xcoord = pdf.x
    pdf.cell(190, 265, '', 1, 1, 'C')
    pdf.y = top
    pdf.image("IITPLOGO2.PNG",x = xcoord+1, y = top+1, w = 186, h = 30)
    pdf.ln(30)

    #----------------------------------------------------------#
    stud_info = [["Roll No: ",rollno,"Name: ",name, "Year of Admission: ",yoa], 
                ["Programme: ",programme,"Course: ",course],"",""]
    line_height = 10
    pdf.x = 15
    top = pdf.y
    pdf.cell(175, 23, '', 1, 1, 'C')
    pdf.y = top
    pdf.set_font('Arial', '', 7)
    for row in stud_info:
        pdf.x = 15
        col_no = 1
        for datum in row:
            if col_no == 1:
                col_width = 19
            if col_no == 2:
                col_width = 43
            if col_no == 3:
                col_width = 13
            if col_no == 4:
                col_width = 43
            if col_no == 5:
                col_width = 28
            if col_no == 6:
                col_width = 13
            pdf.multi_cell(col_width, line_height, datum, border=0, ln=3, max_line_height=pdf.font_size)
            col_no = col_no + 1
        pdf.ln(line_height)

        #----------------------------------------------------------#
    xcoord = pdf.x
    pdf.y = pdf.y-13
    ycoord = pdf.y
    offset = 3
    offset_y = 85
    for key in transcript_data:
        sem_table_list = []
        if key == "5":
            break
        if key == "3" :
            pdf.y = pdf.y + offset_y
            ycoord = pdf.y
            pdf.x = xcoord
            offset = 3

        for subject_wise_data in transcript_data[key]:
            sem_table_list.append(subject_wise_data)
        # print(sem_table_list)
        # break
        pdf.x = pdf.x+offset
        pdf.set_font('Arial',"", 8)
        sem_name = "Semester "+str(key)
        pdf.set_font('Arial', 'BU', 6)
        pdf.cell(20, 10,sem_name, 0, 1, 'C')
        pdf.ln(1)
        for row in sem_table_list[:-1]:
            i = 1
            
            pdf.x = pdf.x+offset
            for datum in row[1:]:
                if i == 1:
                    cell_width = 9
                if i == 2:
                    cell_width = 40
                if i == 3:
                    cell_width = 9
                if i == 4:
                    cell_width = 9
                if i == 5:
                    cell_width = 9
                pdf.set_font('Arial', '', 5)
                
                pdf.multi_cell(cell_width, 7, datum, border=1, ln=3, max_line_height=pdf.font_size)
                i=i+1
            pdf.ln(7)
        i=1
        # print(sem_table_list[-1])
        pdf.x = pdf.x+offset
        pdf.y = pdf.y+2
        pdf.set_font('Arial', 'B', 7)
        overalldata = "Credits Taken: {creditstaken}     Credits Cleared: {creditscleared}   SPI: {spi}  CPI: {cpi}".format(creditstaken=sem_table_list[-1][1],creditscleared = sem_table_list[-1][3],spi=sem_table_list[-1][2],cpi=sem_table_list[-1][4] )
        pdf.cell(77, 8, overalldata, 1, 1, 'C')
        pdf.ln(7)
        

        pdf.y = ycoord
        offset = offset + 90


    today = datetime.datetime.now()
    today = today.strftime("%d %b %Y, %H:%M")
    today = "Date Generated: "+str(today)
    pdf.x = 30
    pdf.y = 240
    pdf.set_font('Arial', '', 9)
    pdf.cell(20, 10,today, 0, 1, 'C')

    pdf.x = 70
    pdf.y = 240
    pdf.image("SEAL.PNG",x = 80, y = 235, w = 35, h = 25)


    pdf.x = 150
    pdf.y = 240
    pdf.cell(20, 10,"Assistant Registrar (Academic)", 0, 1, 'C')
    file_name = str(rollno)+".pdf"
    pdf.output(file_name)



    
try:
    os.mkdir("output")
except:
    pass

func()



