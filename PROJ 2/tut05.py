import csv
from openpyxl import Workbook
import os
def generate_marksheet():
    nameRollMapping = {}
    with open('sample_input/names-roll.csv','r') as namesRollFile:
        reader = csv.DictReader(namesRollFile)
        for row in reader:
            nameRollMapping[row['Roll']] = row['Name']
    subjectMapping = {}
    with open('sample_input/subjects_master.csv','r') as subject_master:
        reader = csv.DictReader(subject_master)
        for row in reader:
            subjectMapping[row['subno']] = [row['subname'],row['crd'],row['ltp']]
    # print(subjectMapping)
    gradeswithsem ={}
    rolls = []
    with open('sample_input/grades.csv','r') as gradesFile:
        reader = csv.DictReader(gradesFile)
        for row in reader:
            if row['Roll'] in rolls:
                gradeswithsem[row['Roll']].append([row['Sem'],row['SubCode'],subjectMapping[row['SubCode']][0],subjectMapping[row['SubCode']][2],row['Credit'],row['Sub_Type'],row['Grade']])
            else:
                gradeswithsem[row['Roll']] = [[row['Sem'],row['SubCode'],subjectMapping[row['SubCode']][0],subjectMapping[row['SubCode']][2],row['Credit'],row['Sub_Type'],row['Grade']]]
                rolls.append(row['Roll'])
    # print(gradeswithsem)
    grade = {'AA': 10,"AB":9,'BB':8,' BB':8,'BC': 7, 'CC':6,'CD':5,'DD':4,'F':0,'I':0,'F*':0,'I*':0,'DD*':4}
    overalldata = {}
    for key in gradeswithsem:
        overalldata[key] = []
    # print(overalldata)
    for key in gradeswithsem:
        #semdata n rows 3 columns semno,semcred,spi,totcred,cpi
        semadded = []
        semdata1 = []
        
        tot_sem_cred = 0
        tot_cred = 0
        sem = 1
        numerator = 0
        semdata = [0,0,0,0,0]
        loopno = 1
        cpi = 0
        for item in gradeswithsem[key]:
            # print(loopno,"  ",len(gradeswithsem[key]))
            if sem == 9:
                break
            if int(item[0]) == sem:
                tot_sem_cred = tot_sem_cred + int(item[4])
                semdata[1] = tot_sem_cred
                numerator = numerator + int(item[4])*grade[item[6]]
            
            else:
                semdata[1] = tot_sem_cred
                spi = numerator/tot_sem_cred
                cpi = (cpi*tot_cred+spi*tot_sem_cred)/(tot_cred+tot_sem_cred)
                semdata[4] = round(cpi,2)
                tot_cred = tot_cred+tot_sem_cred
                semdata[3] = tot_cred
                spi = numerator/tot_sem_cred
                
                semdata[2] = round(spi,2)
                semdata[0] = sem
                semdata1.append(semdata)
                overalldata[key].append(semdata)
                sem = sem + 1
                semdata = [0,0,0,0,0]
                tot_sem_cred = 0
                numerator=0
                tot_sem_cred = tot_sem_cred + int(item[4])
                numerator = int(item[4])*grade[item[6]]
            if loopno == len(gradeswithsem[key]):
                # print("yes")
                semdata[1] = tot_sem_cred
                spi = numerator/tot_sem_cred
                cpi = (cpi*tot_cred+spi*tot_sem_cred)/(tot_cred+tot_sem_cred)
                semdata[4] = round(cpi,2)
                tot_cred = tot_cred+tot_sem_cred
                semdata[3] = tot_cred
                spi = numerator/tot_sem_cred
                semdata[2] = round(spi,2)
                semdata[0] = sem
                semdata1.append(semdata)
                overalldata[key].append(semdata)
            loopno = loopno+1
    
    print(gradeswithsem)
    # print(overalldata)
    return
    for key in gradeswithsem:
        stud_name = key
        file_name = "output/"+ stud_name+".xlsx"
        wb = Workbook()
        Sheet1 = wb.active
        Sheet1.title = "Overall"
        Sheet1.cell(row=1,column=1).value = "Roll No"
        Sheet1.cell(row=1,column=2).value = stud_name
        Sheet1.cell(row=2,column=1).value = "Name of student"
        Sheet1.cell(row=2,column=2).value = nameRollMapping[stud_name]
        Sheet1.cell(row=3,column=1).value = "Discipline"
        Sheet1.cell(row=3,column=2).value = stud_name[4:6]
        Sheet1.cell(row=4,column=1).value = "Semester No"
        Sheet1.cell(row=5,column=1).value = "Semester wise credit taken"
        Sheet1.cell(row=6,column=1).value = "SPI"
        Sheet1.cell(row=7,column=1).value = "Total credits taken"
        Sheet1.cell(row=8,column=1).value = "CPI"
        col_no = 2
        for item in overalldata[stud_name]:
            Sheet1.cell(row=4,column=col_no).value = item[0]
            Sheet1.cell(row=5,column=col_no).value = item[1]
            Sheet1.cell(row=6,column=col_no).value = item[2]
            Sheet1.cell(row=7,column=col_no).value = item[3]
            Sheet1.cell(row=8,column=col_no).value = item[4]
            col_no = col_no + 1

        # print(gradeswithsem[stud_name])
        total_sems = gradeswithsem[stud_name][-1][0]
        # print(total_sems)
        start = 0
        max_start = len(gradeswithsem[stud_name])
        # print(max_start)
        for i in range(int(total_sems)):
            ws1 = wb.create_sheet("Sheet1")
            ws1.title = "sem"+str(i+1)
            ws1.cell(row=1,column=1).value = "Sl No"
            ws1.cell(row=1,column=2).value = "Subject No"
            ws1.cell(row=1,column=3).value = "Subject Name"
            ws1.cell(row=1,column=4).value = "L-T-P"
            ws1.cell(row=1,column=5).value = "Credit"
            ws1.cell(row=1,column=6).value = "Subject Type"
            ws1.cell(row=1,column=7).value = "Grade"
            row_no = 2
            while(int(gradeswithsem[stud_name][start][0]) == (i+1)):
                # print(int(gradeswithsem[stud_name][start][0]))
                # print(start)
                ws1.cell(row=row_no,column=1).value = row_no-1
                ws1.cell(row=row_no,column=2).value = gradeswithsem[stud_name][start][1]
                ws1.cell(row=row_no,column=3).value = gradeswithsem[stud_name][start][2]
                ws1.cell(row=row_no,column=4).value = gradeswithsem[stud_name][start][3]
                ws1.cell(row=row_no,column=5).value = int(gradeswithsem[stud_name][start][4])
                ws1.cell(row=row_no,column=6).value = gradeswithsem[stud_name][start][5]
                ws1.cell(row=row_no,column=7).value = gradeswithsem[stud_name][start][6]
                start = start+1
                row_no = row_no+1
                if start>max_start-1:
                    break
                




            

        wb.save(file_name)
    



    
    return
    
    
try:
    os.mkdir("output")
except:
    pass

generate_marksheet()